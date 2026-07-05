from datetime import date, datetime, time, timedelta
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from src.modules.attendance.service import build_report_workbook


def make_log(emp_id, day, checkin, checkout=None):
    duration = checkout - checkin if checkout else None
    return SimpleNamespace(
        emp_id=emp_id,
        working_date=day,
        checkin_time=checkin,
        checkout_time=checkout,
        working_duration=duration,
    )


def test_report_workbook_summary_and_detail():
    employees = [
        SimpleNamespace(emp_id=1, emp_code="NV001", name="Nguyen A"),
        SimpleNamespace(emp_id=2, emp_code="NV002", name="Tran B"),
    ]
    logs = [
        # on time, 9.5h
        make_log(1, date(2026, 6, 1), datetime(2026, 6, 1, 8, 0), datetime(2026, 6, 1, 17, 30)),
        # late (after 10:00), 8h
        make_log(1, date(2026, 6, 2), datetime(2026, 6, 2, 10, 15), datetime(2026, 6, 2, 18, 15)),
        # open log: no checkout — counts as a day, no hours
        make_log(1, date(2026, 6, 3), datetime(2026, 6, 3, 8, 30)),
    ]

    content = build_report_workbook(employees, logs, check_in_end=time(10, 0))
    wb = load_workbook(BytesIO(content))

    assert wb.sheetnames == ["Summary", "Detail"]

    summary = list(wb["Summary"].values)
    assert summary[0] == ("Employee Code", "Name", "Days Worked", "Total Hours", "Late Count")
    assert summary[1] == ("NV001", "Nguyen A", 3, 17.5, 1)
    # employee with zero logs still appears
    assert summary[2] == ("NV002", "Tran B", 0, 0, 0)

    detail = list(wb["Detail"].values)
    assert len(detail) == 1 + len(logs)
    assert detail[1] == ("NV001", "Nguyen A", "2026-06-01", "08:00:00", "17:30:00", 9.5)
    # open log has empty checkout/hours (openpyxl reads "" back as None)
    assert detail[3][4] is None and detail[3][5] is None
